# for excel_country_financial_filter
# to find the raws, which are in sheet 1 and column 1 same as 'Government'. also, the units sold more than 1000 times
# raw output includes country, units sold, sales, profits, and date
# count how many times Germany appears, if raw[0] is 'Government'

# for excel_professor_rate_filter
# to find the highest quality among all professors

import os
from openpyxl import load_workbook


def counter_financial():
    count = 0
    file_path = os.path.abspath('Financial Sample.xlsx')
    wb = load_workbook(file_path)
    sheet_list = wb.worksheets
    sheet = sheet_list[0]
    for raw in sheet.iter_rows(min_row=2):
        if raw[0].value == 'Government':
            if raw[1].value == 'Germany':
                count += 1
            # need to change datetime formate
            if raw[4].value > 1000:
                format_date = raw[12].value
                date = format_date.strftime('%Y/%m/%d')
                num_dec9 = raw[9].value
                num_9 = round(num_dec9, 2)
                num_dec11 = raw[11].value
                num_11 = round(num_dec11, 2)
                target_list = [raw[1].value, raw[4].value, num_9, num_11, date]
                print(target_list)
    print("Germany counts ", count)


def professor():
    file_path = os.path.abspath('Financial Sample.xlsx')
    wb = load_workbook(file_path)
    sheet_list = wb.worksheets
    sheet = sheet_list[1]
    rate_list = []
    name_list = []
    for raw in sheet.iter_rows(min_row=2):
        n = raw[0].value
        r = raw[1].value
        if r is not None:
            rate_list.append(r)
        if n:
            name_list.append(n)
    high_rate = max(rate_list)
    high_rate_list = [index for index, rate_num in enumerate(rate_list) if rate_num == high_rate]
    for value in high_rate_list:
        print("the professor {} has the highest rate {}".format(name_list[value], high_rate))


if __name__ == '__main__':
    counter_financial()
    print("====================================")
    professor()
